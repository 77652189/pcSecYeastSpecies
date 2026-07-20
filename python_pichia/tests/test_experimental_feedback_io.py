from __future__ import annotations

import csv
import json
from dataclasses import asdict, replace
from enum import Enum

from openpyxl import Workbook
import pytest

from pcsec_pichia.experimental_feedback import (
    ConditionContext,
    ExperimentBundle,
    ExperimentRecord,
    HostContext,
    InterventionRecord,
    InterventionType,
    MeasurementRecord,
    MeasurementStatus,
    SchemaValidationError,
    load_experiment_bundle,
    validate_experiment_bundle,
    write_experiment_feedback_cache,
)


def _bundle() -> ExperimentBundle:
    experiment = ExperimentRecord(
        experiment_id="HLF-IO-R1",
        target_id="hLF",
        host=HostContext("Komagataella phaffii", "X33", "X33"),
        batch_id="B01",
        condition=ConditionContext("BMMY, methanol, shake_flask, 250 rpm", 72.0),
    )
    return ExperimentBundle(
        experiments=(experiment,),
        interventions=(
            InterventionRecord(
                experiment_id=experiment.experiment_id,
                intervention_id="CONTROL-1",
                component_index=1,
                intervention_type=InterventionType.CONTROL,
            ),
        ),
        measurements=(
            MeasurementRecord(
                experiment_id=experiment.experiment_id,
                measurement_id="TITER-T1",
                assay_type="titer",
                assay_method="ELISA",
                compartment="extracellular",
                raw_value=None,
                raw_unit="mg/L",
                canonical_value=None,
                canonical_unit="mg/L",
                status=MeasurementStatus.ASSAY_FAILED,
                technical_replicate_id="T1",
                status_reason="sanitized plate control failure",
            ),
        ),
    )


def test_jsonl_cache_roundtrip_preserves_raw_failure_state(tmp_path) -> None:
    outputs = write_experiment_feedback_cache(_bundle(), tmp_path)

    loaded = load_experiment_bundle(outputs.validated_records_path)

    assert loaded == _bundle()
    assert loaded.measurements[0].raw_value is None
    assert loaded.measurements[0].raw_unit == "mg/L"
    assert loaded.measurements[0].status is MeasurementStatus.ASSAY_FAILED
    manifest = json.loads(outputs.manifest_path.read_text(encoding="utf-8"))
    assert manifest["validated_record_count"] == 3
    assert manifest["conflict_count"] == 0


def test_csv_import_reports_duplicates_conflicts_bad_units_and_missing_conditions(tmp_path) -> None:
    experiment = _bundle().experiments[0]
    conflicting_experiment = ExperimentRecord(
        experiment_id=experiment.experiment_id,
        target_id="OPN",
        host=experiment.host,
        batch_id=experiment.batch_id,
        condition=experiment.condition,
    )
    missing_condition = ExperimentRecord(
        experiment_id="OPN-MISSING-CONTEXT",
        target_id="OPN",
        host=experiment.host,
        batch_id="B02",
        condition=ConditionContext("missing", None),
    )
    bad_unit = MeasurementRecord(
        experiment_id=experiment.experiment_id,
        measurement_id="BAD-UNIT-1",
        assay_type="titer",
        assay_method="ELISA",
        compartment="extracellular",
        raw_value=0.012,
        raw_unit="g/L",
        canonical_value=0.012,
        canonical_unit="g/L",
        status=MeasurementStatus.VALID,
    )
    orphan_measurement = MeasurementRecord(
        experiment_id="MISSING-EXPERIMENT",
        measurement_id="ORPHAN-TITER",
        assay_type="titer",
        assay_method="ELISA",
        compartment="extracellular",
        raw_value=1.0,
        raw_unit="mg/L",
        canonical_value=1.0,
        canonical_unit="mg/L",
        status=MeasurementStatus.VALID,
    )
    rows = [
        ("experiment", experiment),
        ("experiment", experiment),
        ("experiment", conflicting_experiment),
        ("experiment", missing_condition),
        ("intervention", _bundle().interventions[0]),
        ("measurement", bad_unit),
        ("measurement", orphan_measurement),
    ]
    csv_path = tmp_path / "sanitized_import.csv"
    with csv_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=["record_type", "payload_json"])
        writer.writeheader()
        for record_type, record in rows:
            writer.writerow(
                {
                    "record_type": record_type,
                    "payload_json": json.dumps(asdict(record), default=_enum_value),
                }
            )

    bundle = load_experiment_bundle(csv_path)
    validation = validate_experiment_bundle(bundle)
    outputs = write_experiment_feedback_cache(bundle, tmp_path / "cache")

    assert validation.is_valid is False
    assert any(issue.code == "record_id_conflict" for issue in validation.errors)
    assert any(issue.code == "unit_validation_error" for issue in validation.errors)
    assert any(issue.code == "condition_missing" for issue in validation.warnings)
    assert any("duplicate_record_ignored" in warning for warning in bundle.warnings)
    conflicts = [json.loads(line) for line in outputs.conflicts_path.read_text(encoding="utf-8").splitlines()]
    assert any(row["code"] == "record_id_conflict" for row in conflicts)
    assert any(row["code"] == "unit_validation_error" for row in conflicts)
    validated_text = outputs.validated_records_path.read_text(encoding="utf-8")
    assert "BAD-UNIT-1" not in validated_text
    assert "ORPHAN-TITER" not in validated_text
    manifest = json.loads(outputs.manifest_path.read_text(encoding="utf-8"))
    assert manifest["source_file"] == str(csv_path)
    assert len(manifest["source_sha256"]) == 64
    assert manifest["imported_at"]
    assert any("duplicate_record_ignored" in warning for warning in manifest["warnings"])
    assert manifest["validated_record_count"] == 3
    assert manifest["conflict_count"] == 3


def test_xlsx_import_uses_the_same_record_envelope_contract_as_csv(tmp_path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "records"
    worksheet.append(("record_type", "payload_json"))
    bundle = _bundle()
    for record_type, records in (
        ("experiment", bundle.experiments),
        ("intervention", bundle.interventions),
        ("measurement", bundle.measurements),
    ):
        for record in records:
            worksheet.append(
                (record_type, json.dumps(asdict(record), default=_enum_value))
            )
    metadata = workbook.create_sheet("metadata")
    metadata.append(("generated_by", "legacy canonical exporter"))
    metadata.append(("format_version", "2026-01"))
    xlsx_path = tmp_path / "sanitized_import.xlsx"
    workbook.save(xlsx_path)

    loaded = load_experiment_bundle(xlsx_path)
    validation = validate_experiment_bundle(loaded)

    assert validation.is_valid is True
    assert loaded.experiments == bundle.experiments
    assert loaded.interventions == bundle.interventions
    assert loaded.measurements == bundle.measurements
    assert loaded.import_manifest is not None
    assert loaded.import_manifest.source_file == str(xlsx_path)
    assert len(loaded.import_manifest.source_sha256) == 64


def test_corrupt_xlsx_is_reported_as_a_schema_validation_error(tmp_path) -> None:
    xlsx_path = tmp_path / "corrupt.xlsx"
    xlsx_path.write_bytes(b"not-an-xlsx-container")

    with pytest.raises(SchemaValidationError, match="invalid XLSX experiment bundle"):
        load_experiment_bundle(xlsx_path)


def test_incomplete_canonical_envelopes_are_not_misclassified_as_wide_templates(
    tmp_path,
) -> None:
    jsonl_path = tmp_path / "incomplete.jsonl"
    jsonl_path.write_text(
        json.dumps({"record_type": "experiment"}) + "\n",
        encoding="utf-8",
    )
    with pytest.raises(SchemaValidationError, match="record_type and record"):
        load_experiment_bundle(jsonl_path)

    csv_path = tmp_path / "incomplete.csv"
    csv_path.write_text("record_type\nexperiment\n", encoding="utf-8")
    with pytest.raises(SchemaValidationError, match="missing columns: payload_json"):
        load_experiment_bundle(csv_path)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(("payload_json",))
    worksheet.append(("{}",))
    xlsx_path = tmp_path / "incomplete.xlsx"
    workbook.save(xlsx_path)
    with pytest.raises(SchemaValidationError, match="missing columns: record_type"):
        load_experiment_bundle(xlsx_path)


def _enum_value(value: object) -> object:
    if isinstance(value, Enum):
        return value.value
    raise TypeError(type(value).__name__)


def test_measurement_compartment_must_be_canonical() -> None:
    invalid = MeasurementRecord(
        experiment_id="HLF-IO-R1",
        measurement_id="BAD-COMPARTMENT",
        assay_type="titer",
        assay_method="ELISA",
        compartment="secreted",
        raw_value=1.0,
        raw_unit="mg/L",
        canonical_value=1.0,
        canonical_unit="mg/L",
        status=MeasurementStatus.VALID,
    )

    validation = validate_experiment_bundle(
        ExperimentBundle(
            experiments=_bundle().experiments,
            interventions=_bundle().interventions,
            measurements=(invalid,),
        )
    )

    assert validation.is_valid is False
    assert any("compartment" in issue.message for issue in validation.errors)


def test_non_finite_measurement_value_is_rejected() -> None:
    valid = _bundle().measurements[0]
    invalid = replace(
        valid,
        status=MeasurementStatus.VALID,
        status_reason="",
        raw_value=float("nan"),
        canonical_value=float("nan"),
    )

    validation = validate_experiment_bundle(
        ExperimentBundle(
            experiments=_bundle().experiments,
            interventions=_bundle().interventions,
            measurements=(invalid,),
        )
    )

    assert validation.is_valid is False
    assert any("finite" in issue.message for issue in validation.errors)
