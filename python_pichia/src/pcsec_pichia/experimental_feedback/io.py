from __future__ import annotations

import csv
import dataclasses
import hashlib
import json
from dataclasses import asdict, dataclass
from datetime import datetime, timezone
from enum import Enum
from pathlib import Path
from typing import Any, Iterable, Mapping, get_type_hints
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from pcsec_pichia.experimental_feedback.quality import validate_experiment_bundle
from pcsec_pichia.experimental_feedback.fermentation_template import (
    FERMENTATION_TEMPLATE_ADAPTER_ID,
    map_fermentation_template_rows,
    merge_fermentation_template_metadata,
)
from pcsec_pichia.experimental_feedback.schema import (
    SCHEMA_VERSION,
    ConditionContext,
    ExperimentBundle,
    ExperimentImportConflict,
    ExperimentImportManifest,
    ExperimentRecord,
    ExperimentalFeedbackError,
    HostContext,
    InterventionRecord,
    MeasurementRecord,
    PredictionLinkRecord,
    SchemaValidationError,
)


VALIDATED_RECORDS_FILENAME = "validated_records.jsonl"
CONFLICTS_FILENAME = "conflicts.jsonl"
WARNINGS_FILENAME = "warnings.jsonl"
MANIFEST_FILENAME = "manifest.json"


@dataclass(frozen=True)
class ExperimentFeedbackOutputs:
    validated_records_path: Path
    conflicts_path: Path
    warnings_path: Path
    manifest_path: Path


@dataclass(frozen=True)
class _LoadedExperimentRecords:
    records: tuple[tuple[str, object], ...]
    adapter_id: str
    warnings: tuple[str, ...] = ()
    metadata: Mapping[str, object] | None = None
    source_record_count: int | None = None


def load_experiment_bundle(
    path: str | Path,
    *,
    metadata: Mapping[str, object] | None = None,
) -> ExperimentBundle:
    resolved = Path(path)
    if resolved.suffix.lower() == ".csv":
        loaded = _load_csv_records(resolved, metadata=metadata)
    elif resolved.suffix.lower() == ".xlsx":
        loaded = _load_xlsx_records(resolved, metadata=metadata)
    elif resolved.suffix.lower() == ".jsonl":
        loaded = _load_jsonl_records(resolved, metadata=metadata)
    else:
        raise SchemaValidationError(f"unsupported experiment bundle format: {resolved.suffix}")
    unique_records, conflicts, dedupe_warnings = _dedupe_records(loaded.records)
    warnings = tuple(dict.fromkeys((*loaded.warnings, *dedupe_warnings)))
    manifest = ExperimentImportManifest(
        source_file=str(resolved),
        source_sha256=hashlib.sha256(resolved.read_bytes()).hexdigest(),
        imported_at=datetime.now(timezone.utc).isoformat(),
        record_count=(
            loaded.source_record_count
            if loaded.source_record_count is not None
            else len(loaded.records)
        ),
        adapter_id=loaded.adapter_id,
        metadata_json=json.dumps(
            _json_ready(loaded.metadata or {}), ensure_ascii=False, sort_keys=True
        ),
        warnings=warnings,
    )
    manifest.validate()
    return _bundle_from_records(
        unique_records,
        source_file=str(resolved),
        import_manifest=manifest,
        import_conflicts=tuple(conflicts),
        warnings=tuple(warnings),
    )


def _load_jsonl_records(
    resolved: Path,
    *,
    metadata: Mapping[str, object] | None,
) -> _LoadedExperimentRecords:
    payloads: list[tuple[int, Mapping[str, object]]] = []
    with resolved.open(encoding="utf-8-sig") as handle:
        for line_number, line in enumerate(handle, start=1):
            if not line.strip():
                continue
            try:
                payload = json.loads(line)
            except json.JSONDecodeError as exc:
                raise SchemaValidationError(f"invalid JSONL at line {line_number}: {exc.msg}") from exc
            if not isinstance(payload, Mapping):
                raise SchemaValidationError(f"JSONL line {line_number} must be an object.")
            payloads.append((line_number, payload))
    envelope_markers = tuple(
        ("record_type" in payload, "record" in payload) for _, payload in payloads
    )
    if any(any(markers) for markers in envelope_markers):
        if not all(all(markers) for markers in envelope_markers):
            raise SchemaValidationError(
                "JSONL canonical envelope requires record_type and record on every data line."
            )
        return _LoadedExperimentRecords(
            records=tuple(
                _record_from_envelope(payload, line_number=line_number)
                for line_number, payload in payloads
            ),
            adapter_id="pcsec_pichia.canonical_envelope.v1",
            source_record_count=len(payloads),
        )
    imported = map_fermentation_template_rows(payloads, metadata=metadata)
    return _LoadedExperimentRecords(
        records=imported.records,
        adapter_id=FERMENTATION_TEMPLATE_ADAPTER_ID,
        warnings=imported.warnings,
        metadata=imported.metadata,
        source_record_count=len(payloads),
    )


def _load_csv_records(
    resolved: Path,
    *,
    metadata: Mapping[str, object] | None,
) -> _LoadedExperimentRecords:
    with resolved.open(encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        if not reader.fieldnames:
            raise SchemaValidationError("CSV experiment bundle requires a header row.")
        rows = [(line_number, dict(row)) for line_number, row in enumerate(reader, start=2)]
        is_envelope = _tabular_envelope_headers(reader.fieldnames, format_name="CSV")
        if not is_envelope:
            imported = map_fermentation_template_rows(rows, metadata=metadata)
            return _LoadedExperimentRecords(
                records=imported.records,
                adapter_id=FERMENTATION_TEMPLATE_ADAPTER_ID,
                warnings=imported.warnings,
                metadata=imported.metadata,
                source_record_count=len(rows),
            )
        records: list[tuple[str, object]] = []
        for line_number, row in rows:
            try:
                record_payload = json.loads(row.get("payload_json") or "")
            except json.JSONDecodeError as exc:
                raise SchemaValidationError(f"invalid CSV payload_json at line {line_number}: {exc.msg}") from exc
            records.append(
                _record_from_envelope(
                    {"record_type": row.get("record_type"), "record": record_payload},
                    line_number=line_number,
                )
            )
    return _LoadedExperimentRecords(
        records=tuple(records),
        adapter_id="pcsec_pichia.canonical_envelope.v1",
        source_record_count=len(rows),
    )


def _load_xlsx_records(
    resolved: Path,
    *,
    metadata: Mapping[str, object] | None,
) -> _LoadedExperimentRecords:
    try:
        workbook = load_workbook(resolved, read_only=True, data_only=True)
    except (BadZipFile, InvalidFileException, OSError, ValueError) as exc:
        raise SchemaValidationError(f"invalid XLSX experiment bundle: {exc}") from exc
    try:
        worksheet = workbook["records"] if "records" in workbook.sheetnames else next(
            (sheet for sheet in workbook.worksheets if sheet.title != "metadata"),
            workbook.active,
        )
        rows = worksheet.iter_rows(values_only=True)
        header_row = next(rows, None)
        headers = [str(value or "").strip() for value in (header_row or ())]
        required = {"record_type", "payload_json"}
        data_rows = [
            (line_number, row)
            for line_number, row in enumerate(rows, start=2)
            if row and any(value not in (None, "") for value in row)
        ]
        is_envelope = _tabular_envelope_headers(headers, format_name="XLSX")
        if not is_envelope:
            workbook_metadata = _xlsx_metadata(workbook)
            combined_metadata = merge_fermentation_template_metadata(
                workbook_metadata, metadata
            )
            mapped_rows = (
                (
                    line_number,
                    {
                        header: (row[index] if index < len(row) else None)
                        for index, header in enumerate(headers)
                        if header
                    },
                )
                for line_number, row in data_rows
            )
            imported = map_fermentation_template_rows(
                mapped_rows,
                metadata=combined_metadata,
                source_sheet=worksheet.title,
            )
            return _LoadedExperimentRecords(
                records=imported.records,
                adapter_id=FERMENTATION_TEMPLATE_ADAPTER_ID,
                warnings=imported.warnings,
                metadata=imported.metadata,
                source_record_count=len(data_rows),
            )
        record_type_index = headers.index("record_type")
        payload_index = headers.index("payload_json")
        records: list[tuple[str, object]] = []
        for line_number, row in data_rows:
            record_type = row[record_type_index] if record_type_index < len(row) else None
            payload_json = row[payload_index] if payload_index < len(row) else None
            try:
                record_payload = json.loads(str(payload_json or ""))
            except json.JSONDecodeError as exc:
                raise SchemaValidationError(
                    f"invalid XLSX payload_json at row {line_number}: {exc.msg}"
                ) from exc
            records.append(
                _record_from_envelope(
                    {"record_type": record_type, "record": record_payload},
                    line_number=line_number,
                )
            )
        return _LoadedExperimentRecords(
            records=tuple(records),
            adapter_id="pcsec_pichia.canonical_envelope.v1",
            source_record_count=len(data_rows),
        )
    finally:
        workbook.close()


def write_experiment_feedback_cache(
    bundle: ExperimentBundle,
    output_dir: str | Path,
) -> ExperimentFeedbackOutputs:
    validation = validate_experiment_bundle(bundle)
    resolved = Path(output_dir)
    resolved.mkdir(parents=True, exist_ok=True)
    outputs = ExperimentFeedbackOutputs(
        validated_records_path=resolved / VALIDATED_RECORDS_FILENAME,
        conflicts_path=resolved / CONFLICTS_FILENAME,
        warnings_path=resolved / WARNINGS_FILENAME,
        manifest_path=resolved / MANIFEST_FILENAME,
    )
    records = _valid_records(bundle)
    _write_jsonl(
        (
            {"record_type": record_type, "record": _json_ready(asdict(record))}
            for record_type, record in records
        ),
        outputs.validated_records_path,
    )
    conflict_rows = [asdict(conflict) for conflict in bundle.import_conflicts]
    conflict_keys = {(conflict.code, conflict.record_type, conflict.record_id) for conflict in bundle.import_conflicts}
    conflict_rows.extend(
        {
            "code": issue.code,
            "record_type": issue.record_type,
            "record_id": issue.record_id,
            "message": issue.message,
        }
        for issue in validation.errors
        if (issue.code, issue.record_type, issue.record_id) not in conflict_keys
    )
    _write_jsonl(conflict_rows, outputs.conflicts_path)
    _write_jsonl(
        (
            {
                "code": warning.code,
                "record_type": warning.record_type,
                "record_id": warning.record_id,
                "message": warning.message,
            }
            for warning in validation.warnings
        ),
        outputs.warnings_path,
    )
    outputs.manifest_path.write_text(
        json.dumps(
            {
                "schema_version": SCHEMA_VERSION,
                "generated_at": datetime.now(timezone.utc).isoformat(),
                "imported_at": (
                    bundle.import_manifest.imported_at if bundle.import_manifest else ""
                ),
                "source_file": bundle.source_file,
                "source_sha256": (
                    bundle.import_manifest.source_sha256 if bundle.import_manifest else ""
                ),
                "adapter_id": (
                    bundle.import_manifest.adapter_id if bundle.import_manifest else ""
                ),
                "import_metadata": (
                    json.loads(bundle.import_manifest.metadata_json)
                    if bundle.import_manifest
                    else {}
                ),
                "validated_record_count": len(records),
                "conflict_count": len(conflict_rows),
                "warning_count": len(validation.warnings),
                "warnings": list(bundle.import_manifest.warnings) if bundle.import_manifest else [],
                "files": {
                    "validated_records": outputs.validated_records_path.name,
                    "conflicts": outputs.conflicts_path.name,
                    "warnings": outputs.warnings_path.name,
                },
            },
            ensure_ascii=False,
            indent=2,
            sort_keys=True,
        )
        + "\n",
        encoding="utf-8",
    )
    return outputs


def _iter_bundle_records(bundle: ExperimentBundle) -> Iterable[tuple[str, object]]:
    for record in bundle.experiments:
        yield "experiment", record
    for record in bundle.interventions:
        yield "intervention", record
    for record in bundle.measurements:
        yield "measurement", record
    for record in bundle.prediction_links:
        yield "prediction_link", record


def _valid_records(bundle: ExperimentBundle) -> tuple[tuple[str, object], ...]:
    valid: list[tuple[str, object]] = []
    valid_experiment_ids: set[str] = set()
    for record in bundle.experiments:
        try:
            record.validate()
        except ExperimentalFeedbackError:
            continue
        valid.append(("experiment", record))
        valid_experiment_ids.add(record.experiment_id)
    valid_interventions: set[tuple[str, str]] = set()
    for record in bundle.interventions:
        try:
            record.validate()
        except ExperimentalFeedbackError:
            continue
        if record.experiment_id not in valid_experiment_ids:
            continue
        valid.append(("intervention", record))
        valid_interventions.add((record.experiment_id, record.intervention_id))
    for record in bundle.measurements:
        try:
            record.validate()
        except ExperimentalFeedbackError:
            continue
        if record.experiment_id in valid_experiment_ids:
            valid.append(("measurement", record))
    for record in bundle.prediction_links:
        try:
            record.validate()
        except ExperimentalFeedbackError:
            continue
        if (record.experiment_id, record.intervention_id) in valid_interventions:
            valid.append(("prediction_link", record))
    return tuple(valid)


def _dedupe_records(
    records: Iterable[tuple[str, object]],
) -> tuple[list[tuple[str, object]], list[ExperimentImportConflict], list[str]]:
    unique: list[tuple[str, object]] = []
    seen: dict[tuple[str, str], tuple[object, str, str]] = {}
    conflicts: list[ExperimentImportConflict] = []
    warnings: list[str] = []
    for record_type, record in records:
        record_id = _record_id(record_type, record)
        key = (record_type, record_id)
        payload_json = json.dumps(
            _json_ready(asdict(record)), ensure_ascii=False, sort_keys=True, separators=(",", ":")
        )
        comparison_payload_json = json.dumps(
            _dedupe_payload(record_type, record),
            ensure_ascii=False,
            sort_keys=True,
            separators=(",", ":"),
        )
        previous = seen.get(key)
        if previous is None:
            seen[key] = (record, payload_json, comparison_payload_json)
            unique.append((record_type, record))
            continue
        if previous[2] == comparison_payload_json:
            warnings.append(f"duplicate_record_ignored:{record_type}:{record_id}")
            continue
        conflicts.append(
            ExperimentImportConflict(
                code="record_id_conflict",
                record_type=record_type,
                record_id=record_id,
                first_payload_json=previous[1],
                conflicting_payload_json=payload_json,
            )
        )
    return unique, conflicts, warnings


def _dedupe_payload(record_type: str, record: object) -> object:
    payload = _json_ready(asdict(record))
    if record_type == "measurement" and isinstance(payload, dict):
        payload.pop("source_row_number", None)
        payload.pop("source_sheet", None)
    return payload


def _record_id(record_type: str, record: object) -> str:
    if record_type == "experiment":
        return str(record.experiment_id)  # type: ignore[attr-defined]
    if record_type == "intervention":
        return f"{record.experiment_id}/{record.intervention_id}"  # type: ignore[attr-defined]
    if record_type == "measurement":
        return f"{record.experiment_id}/{record.measurement_id}"  # type: ignore[attr-defined]
    if record_type == "prediction_link":
        return "/".join(
            (
                str(record.experiment_id),  # type: ignore[attr-defined]
                str(record.intervention_id),  # type: ignore[attr-defined]
                str(record.prediction_run_id),  # type: ignore[attr-defined]
                str(record.evidence_id),  # type: ignore[attr-defined]
            )
        )
    raise SchemaValidationError(f"unsupported record_type: {record_type}")


def _record_from_envelope(payload: object, *, line_number: int) -> tuple[str, object]:
    if not isinstance(payload, Mapping):
        raise SchemaValidationError(f"JSONL line {line_number} must be an object.")
    record_type = str(payload.get("record_type") or "")
    record = payload.get("record")
    if not isinstance(record, Mapping):
        raise SchemaValidationError(f"JSONL line {line_number} record must be an object.")
    try:
        return record_type, _record_from_dict(record_type, record)
    except (TypeError, ValueError, KeyError) as exc:
        raise SchemaValidationError(
            f"invalid {record_type or 'unknown'} record at line {line_number}: {exc}"
        ) from exc


def _coerce_enum_fields(cls: type, values: dict[str, Any]) -> dict[str, Any]:
    """Convert any enum-typed field present in `values` from its serialized value back
    to the enum member, by introspecting `cls`'s own annotations rather than naming each
    field. New enum-typed fields on these dataclasses are covered automatically."""
    hints = get_type_hints(cls)
    for field in dataclasses.fields(cls):
        hint = hints.get(field.name)
        if not isinstance(hint, type) or not issubclass(hint, Enum):
            continue
        if field.name in values and not isinstance(values[field.name], hint):
            values[field.name] = hint(values[field.name])
    return values


def _record_from_dict(record_type: str, payload: Mapping[str, Any]) -> object:
    values = dict(payload)
    if record_type == "experiment":
        values["host"] = HostContext(**values["host"])
        values["condition"] = ConditionContext(**values["condition"])
        values = _coerce_enum_fields(ExperimentRecord, values)
        return ExperimentRecord(**values)
    if record_type == "intervention":
        values["warnings"] = tuple(values.get("warnings") or ())
        values = _coerce_enum_fields(InterventionRecord, values)
        return InterventionRecord(**values)
    if record_type == "measurement":
        values = _coerce_enum_fields(MeasurementRecord, values)
        return MeasurementRecord(**values)
    if record_type == "prediction_link":
        values = _coerce_enum_fields(PredictionLinkRecord, values)
        return PredictionLinkRecord(**values)
    raise SchemaValidationError(f"unsupported record_type: {record_type}")


def _xlsx_metadata(workbook: object) -> dict[str, object]:
    if "metadata" not in workbook.sheetnames:  # type: ignore[attr-defined]
        return {}
    worksheet = workbook["metadata"]  # type: ignore[index]
    values: dict[str, object] = {}
    for row_number, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
        key = str(row[0] or "").strip() if row else ""
        if not key:
            continue
        value = row[1] if len(row) > 1 else None
        if row_number == 1 and key.casefold() in {"key", "field", "字段"}:
            value_header = str(value or "").strip().casefold()
            if value_header in {"value", "值"}:
                continue
        if key in values and values[key] not in (None, "") and value not in (None, ""):
            if str(values[key]).strip() != str(value).strip():
                raise SchemaValidationError(
                    f"fermentation import metadata conflict in XLSX sheet: {key}"
                )
        elif value not in (None, "") or key not in values:
            values[key] = value
    return values


def _tabular_envelope_headers(
    headers: Iterable[str],
    *,
    format_name: str,
) -> bool:
    required = {"record_type", "payload_json"}
    present = required.intersection(str(item or "").strip() for item in headers)
    if present and present != required:
        missing = ", ".join(sorted(required - present))
        raise SchemaValidationError(
            f"{format_name} canonical envelope is incomplete; missing columns: {missing}."
        )
    return present == required


def _bundle_from_records(
    records: Iterable[tuple[str, object]],
    *,
    source_file: str = "",
    import_manifest: ExperimentImportManifest | None = None,
    import_conflicts: tuple[ExperimentImportConflict, ...] = (),
    warnings: tuple[str, ...] = (),
) -> ExperimentBundle:
    by_type: dict[str, list[object]] = {
        "experiment": [],
        "intervention": [],
        "measurement": [],
        "prediction_link": [],
    }
    for record_type, record in records:
        by_type[record_type].append(record)
    bundle = ExperimentBundle(
        experiments=tuple(by_type["experiment"]),  # type: ignore[arg-type]
        interventions=tuple(by_type["intervention"]),  # type: ignore[arg-type]
        measurements=tuple(by_type["measurement"]),  # type: ignore[arg-type]
        prediction_links=tuple(by_type["prediction_link"]),  # type: ignore[arg-type]
        source_file=source_file,
        warnings=warnings,
        import_manifest=import_manifest,
        import_conflicts=import_conflicts,
    )
    return bundle


def _write_jsonl(rows: Iterable[Mapping[str, Any]], path: Path) -> None:
    with path.open("w", encoding="utf-8", newline="\n") as handle:
        for row in rows:
            handle.write(json.dumps(_json_ready(row), ensure_ascii=False, sort_keys=True))
            handle.write("\n")


def _json_ready(value: object) -> object:
    if isinstance(value, Enum):
        return value.value
    if isinstance(value, tuple):
        return [_json_ready(item) for item in value]
    if isinstance(value, list):
        return [_json_ready(item) for item in value]
    if isinstance(value, Mapping):
        return {str(key): _json_ready(item) for key, item in value.items()}
    return value


__all__ = [
    "CONFLICTS_FILENAME",
    "ExperimentFeedbackOutputs",
    "MANIFEST_FILENAME",
    "VALIDATED_RECORDS_FILENAME",
    "WARNINGS_FILENAME",
    "load_experiment_bundle",
    "write_experiment_feedback_cache",
]
