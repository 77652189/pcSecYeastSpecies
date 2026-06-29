from __future__ import annotations

import math
from collections import Counter
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook


@dataclass(frozen=True)
class AminoAcidStoichiometry:
    amino_acids: tuple[str, ...]
    translation_substrates: dict[str, str]
    translation_products: dict[str, str]
    degradation_products: dict[str, str]
    energy_substrates: tuple[str, str, str]
    energy_products: tuple[str, str, str, str]

    @classmethod
    def from_workbook(cls, path: Path) -> "AminoAcidStoichiometry":
        workbook = load_workbook(path, data_only=True, read_only=True)
        cytoplasm_rows = list(workbook["cytoplasm"].iter_rows(values_only=True))[1:]
        energy_rows = list(workbook["energy"].iter_rows(values_only=True))[1:]
        amino_acids: list[str] = []
        translation_substrates: dict[str, str] = {}
        translation_products: dict[str, str] = {}
        degradation_products: dict[str, str] = {}
        for row in cytoplasm_rows:
            aa = _text(row[0])
            if not aa:
                continue
            amino_acids.append(aa)
            translation_substrates[aa] = _text(row[2])
            translation_products[aa] = _text(row[4])
            degradation_products[aa] = _text(row[6])
        energy_substrates = tuple(_text(row[2]) for row in energy_rows[:3])
        energy_products = tuple(_text(row[4]) for row in energy_rows[:4])
        if len(energy_substrates) != 3 or len(energy_products) != 4:
            raise ValueError(f"Unexpected energy sheet shape in {path}")
        return cls(
            amino_acids=tuple(amino_acids),
            translation_substrates=translation_substrates,
            translation_products=translation_products,
            degradation_products=degradation_products,
            energy_substrates=energy_substrates,
            energy_products=energy_products,
        )

    def translation_stoichiometry(self, protein_id: str, sequence: str) -> dict[str, float]:
        counts = _aa_counts_with_initiator_methionine(sequence, self.amino_acids)
        stoichiometry: dict[str, float] = {}
        for aa, count in counts.items():
            if count == 0:
                continue
            _accumulate(stoichiometry, self.translation_substrates[aa], -count)
            _accumulate(stoichiometry, self.translation_products[aa], count)
        length = len(sequence)
        total_atp = 2 * length + 3
        total_gtp = 2 * length + 3
        total_h2o = total_atp + total_gtp
        _accumulate(stoichiometry, self.energy_substrates[0], -total_h2o)
        _accumulate(stoichiometry, self.energy_substrates[1], -total_atp)
        _accumulate(stoichiometry, self.energy_substrates[2], -total_gtp)
        _accumulate(stoichiometry, self.energy_products[0], total_h2o)
        _accumulate(stoichiometry, self.energy_products[1], total_atp)
        _accumulate(stoichiometry, self.energy_products[2], total_gtp)
        _accumulate(stoichiometry, self.energy_products[3], total_h2o)
        _accumulate(stoichiometry, f"{protein_id}_peptide[c]", 1.0)
        return stoichiometry

    def signal_peptide_degradation_stoichiometry(
        self,
        protein_id: str,
        signal_peptide_sequence: str,
    ) -> dict[str, float]:
        return self._degradation_stoichiometry(
            sequence=signal_peptide_sequence,
            terminal_metabolite=f"{protein_id}_sp[c]",
            include_initiator_methionine=False,
        )

    def subunit_degradation_stoichiometry(
        self,
        protein_id: str,
        sequence: str,
    ) -> dict[str, float]:
        return self._degradation_stoichiometry(
            sequence=sequence,
            terminal_metabolite=f"{protein_id}_subunit[c]",
            include_initiator_methionine=True,
        )

    def _degradation_stoichiometry(
        self,
        sequence: str,
        terminal_metabolite: str,
        include_initiator_methionine: bool,
    ) -> dict[str, float]:
        counts = _aa_counts_with_initiator_methionine(sequence, self.amino_acids) if include_initiator_methionine else Counter(sequence)
        stoichiometry: dict[str, float] = {}
        for aa in self.amino_acids:
            count = counts.get(aa, 0)
            if count:
                _accumulate(stoichiometry, self.degradation_products[aa], count)
        energy_count = math.floor(1.3 * len(sequence))
        _accumulate(stoichiometry, self.energy_substrates[0], -energy_count)
        _accumulate(stoichiometry, self.energy_substrates[1], -energy_count)
        _accumulate(stoichiometry, self.energy_products[0], energy_count)
        _accumulate(stoichiometry, self.energy_products[1], energy_count)
        _accumulate(stoichiometry, self.energy_products[3], energy_count)
        _accumulate(stoichiometry, terminal_metabolite, -1.0)
        return stoichiometry


def _aa_counts_with_initiator_methionine(sequence: str, amino_acids: tuple[str, ...]) -> Counter[str]:
    counts = Counter(sequence)
    if "M" in amino_acids:
        counts["M"] += 1
    return counts


def _accumulate(stoichiometry: dict[str, float], metabolite_id: str, coefficient: float) -> None:
    if not metabolite_id or coefficient == 0:
        return
    stoichiometry[metabolite_id] = stoichiometry.get(metabolite_id, 0.0) + float(coefficient)
    if stoichiometry[metabolite_id] == 0:
        del stoichiometry[metabolite_id]


def _text(value: object) -> str:
    return str(value or "").strip()
